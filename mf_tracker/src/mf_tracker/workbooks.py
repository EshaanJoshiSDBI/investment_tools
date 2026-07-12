from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import polars as pl

from .errors import UnsupportedWorkbookError, WorkbookReadError

OOXML_SIGNATURE = b"PK\x03\x04"
OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")


@dataclass(slots=True)
class RawWorkbook:
    reader: str
    sheets: dict[str, pl.DataFrame]


def detect_workbook_format(path: str | Path) -> str:
    source = Path(path)
    with source.open("rb") as handle:
        signature = handle.read(8)
    if signature.startswith(OOXML_SIGNATURE):
        return "ooxml"
    if signature.startswith(OLE_SIGNATURE):
        return "xls"
    raise UnsupportedWorkbookError(
        f"Unsupported workbook signature for {source.name}: {signature.hex()}"
    )


def _rows_to_frame(rows: Iterable[Iterable[Any]]) -> pl.DataFrame:
    materialized = [list(row) for row in rows]
    while materialized and not any(value not in (None, "") for value in materialized[-1]):
        materialized.pop()
    if not materialized:
        return pl.DataFrame({"source_row": []}, schema={"source_row": pl.Int64})
    width = max(len(row) for row in materialized)
    # Ignore formatting-only phantom columns but retain all columns containing values.
    last_used = 0
    for row in materialized:
        for index, value in enumerate(row):
            if value not in (None, ""):
                last_used = max(last_used, index + 1)
    width = max(last_used, 1)
    records: dict[str, list[Any]] = {"source_row": []}
    for index in range(width):
        records[f"column_{index + 1}"] = []
    for row_number, row in enumerate(materialized, start=1):
        records["source_row"].append(row_number)
        for index in range(width):
            records[f"column_{index + 1}"].append(row[index] if index < len(row) else None)
    # Mixed spreadsheet cells must enter Polars as generic strings/numbers without
    # accidental schema failures; strict=False preserves native numerics where possible.
    return pl.DataFrame(records, strict=False, infer_schema_length=None)


def _read_ooxml(path: Path) -> RawWorkbook:
    from openpyxl import load_workbook

    try:
        with path.open("rb") as handle:
            workbook = load_workbook(handle, read_only=True, data_only=True)
            sheets = {
                sheet.title: _rows_to_frame(sheet.iter_rows(values_only=True))
                for sheet in workbook.worksheets
            }
            workbook.close()
    except Exception as exc:
        raise WorkbookReadError(f"openpyxl could not read {path.name}: {exc}") from exc
    return RawWorkbook(reader="openpyxl", sheets=sheets)


def _read_xls(path: Path) -> RawWorkbook:
    import xlrd

    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
        sheets: dict[str, pl.DataFrame] = {}
        for sheet in workbook.sheets():
            rows = (sheet.row_values(index) for index in range(sheet.nrows))
            sheets[sheet.name] = _rows_to_frame(rows)
        workbook.release_resources()
    except Exception as exc:
        raise WorkbookReadError(f"xlrd could not read {path.name}: {exc}") from exc
    return RawWorkbook(reader="xlrd", sheets=sheets)


def read_workbook(path: str | Path) -> RawWorkbook:
    source = Path(path)
    workbook_format = detect_workbook_format(source)
    return _read_ooxml(source) if workbook_format == "ooxml" else _read_xls(source)

